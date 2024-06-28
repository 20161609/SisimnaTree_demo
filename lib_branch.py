from openpyxl import load_workbook
from collections import deque
import json, os

file_path = 'Tree.xlsx'


class Branch:
    def __init__(self, name):
        self.name = name
        self.children = {}
        self.parent = None
        self.path = 'Home'

    def add_child(self, child):
        child.parent = self
        child.path = f'{self.path}/{child.name}'
        self.children[child.name] = child

    def print_tree(self, prefix=""):
        # This is the root node
        node = self
        if prefix == "":
            print(node.name)
        else:
            print(prefix[:-3] + '└── ' + node.name)

        # If this node has children, recursively call this function on the children
        if node.children:
            count = len(node.children)
            for i, child in enumerate(node.children):
                if i < count - 1:
                    # For all children except the last, extend the prefix by "│   "
                    print_tree(node.children[child], prefix + "│   ")
                else:
                    # For the last child, no vertical line is needed after this point
                    print_tree(node.children[child], prefix + "    ")

    def to_dict(self):
        return {
            self.name: {
                c_name: child.to_dict()[c_name] for c_name, child in self.children.items()
            }
        }


def build_tree():
    root = Branch("Home")
    last_nodes = {0: root}  # Dictionary to keep track of the last node at each level
    tree = load_workbook(filename=file_path)['Budget-Tree']
    for row in tree.iter_rows(values_only=True):
        for level, value in enumerate(row):
            if value:
                new_node = Branch(value.strip())
                last_nodes[level].add_child(new_node)
                last_nodes[level + 1] = new_node  # Update the last node for the current level

                # Remove any nodes that are deeper than the current level since we have moved to a new parent
                for deeper_level in range(level + 2, max(last_nodes.keys()) + 1):
                    last_nodes.pop(deeper_level, None)

    return root


def is_sub_path(sub_path, main_path):
    sub_parts = sub_path.strip('/').split('/')
    main_parts = main_path.strip('/').split('/')
    if len(main_parts) > len(sub_parts):
        return False

    return sub_parts[:len(main_parts)] == main_parts


def load_tree():
    file_path = 'BudgetTree.json'
    initial_data = {"Home": {}}

    if not os.path.exists(file_path):
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(initial_data, f, ensure_ascii=False, indent=4)

    with open(file_path, 'r', encoding='utf-8') as f:
        json_tree = json.load(f)

    return json_tree


def build_tree_from_json(json_tree):
    try:
        root_name = list(json_tree.keys())[0]  # Assuming 'Home' is the root node's name.
        root = Branch(root_name)

        def add_branches(node, data):
            for name, sub_data in data.items():
                child = Branch(name)
                node.add_child(child)
                add_branches(child, sub_data)

        add_branches(root, json_tree[root_name])
        return root
    except KeyError as e:
        print(f"Error building tree from JSON: missing key {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def print_tree(node, prefix=""):
    # This is the root node
    if prefix == "":
        print(node.name)
    else:
        print(prefix[:-3] + '└── ' + node.name)

    # If this node has children, recursively call this function on the children
    if node.children:
        count = len(node.children)
        for i, child in enumerate(node.children):
            if i < count - 1:
                # For all children except the last, extend the prefix by "│   "
                print_tree(node.children[child], prefix + "│   ")
            else:
                # For the last child, no vertical line is needed after this point
                print_tree(node.children[child], prefix + "    ")


def search_branch(branch: Branch, path: str):
    for branchName in path.split('/'):
        if branchName.isnumeric():
            num = int(branchName) - 1
            if 0 <= num < len(branch.children):
                branchName = list(branch.children)[num]
            else:
                return None
        if branchName in branch.children:
            branch = branch.children[branchName]
        else:  # Not valid path
            return None
    return branch


def save_tree(json_tree: dict):
    with open("BudgetTree.json", "w", encoding="utf-8") as f:
        json.dump(json_tree, f)


def get_paths():
    with open('BudgetTree.json', encoding='utf-8') as f:
        root = json.load(f)
    path_box = []

    def dfs(node, path_list):
        path_txt = '/'.join(path_list)
        if len(path_txt) > 0:
            path_box.append(path_txt)
        for child_name in node:
            path_list.append(child_name)
            dfs(node[child_name], path_list)
            path_list.pop()

    dfs(root, deque([]))
    return path_box
