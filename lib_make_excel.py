from datetime import datetime, timedelta
from lib_branch import Branch
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from lib_sanitizer import format_cost
from collections import deque

import os
import lib_sanitizer as Df
import pandas as pd


class AccountBookGenerator:
    def __init__(self, data, headers):
        self.count = 0
        self.headers = headers
        self.data_for_excel = []
        self.reformat_data(data)

    def reformat_data(self, data):
        total_in, total_out = 0, 0
        if len(self.headers) == 6:  # Daily Report
            for row in data:
                self.count += 1
                r_date, r_branch, r_cashflow, r_desc = row
                r_date = r_date.split()[0]

                _in = Df.format_cost(r_cashflow) if r_cashflow > 0 else '-'
                _out = Df.format_cost(-r_cashflow) if r_cashflow <= 0 else '-'

                total_in += max(r_cashflow, 0)
                total_out += -min(r_cashflow, 0)

                balance = total_in - total_out
                new_row = (r_date, r_branch, _in, _out, Df.format_cost(balance), r_desc)
                self.data_for_excel.append(new_row)
            balance = total_in - total_out
            total_row = ('Total', '', Df.format_cost(total_in), Df.format_cost(total_out), Df.format_cost(balance), '')
            self.data_for_excel.append(total_row)
        elif len(self.headers) == 4:  # Monthly Report
            for row in data:
                monthly, _in, _out = row
                total_in += _in
                total_out += -_out
                balance = (total_in - total_out)

                _in = Df.format_cost(_in) if _in != 0 else '-'
                _out = Df.format_cost(-_out) if _out != 0 else '-'

                new_row = (monthly, _in, _out, Df.format_cost(balance))
                self.data_for_excel.append(new_row)

            balance = total_in - total_out
            total_row = ('Total', Df.format_cost(total_in), Df.format_cost(total_out), Df.format_cost(balance))
            self.data_for_excel.append(total_row)

    def make_excel(self, file_path):  # 엑셀 파일 출력
        try:
            df = pd.DataFrame(self.data_for_excel, columns=self.headers)
            wb = Workbook()
            ws = wb.active
            width_size = [12] * len(self.headers)  # [12, 12, 12, 12, 12, 12]

            # add data frame onto Excel worksheet.
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # styling headers
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")

                    # styling total row
                    if r_idx == len(df):
                        cell.font = Font(underline="single")
                        cell.border = Border(bottom=Side(style='thin'))

                    # modify cell's width
                    width_size[c_idx] = max(len(str(value)) * 1.5, width_size[c_idx])
                    column_width = width_size[c_idx]
                    ws.column_dimensions[cell.column_letter].width = column_width

            # save file
            wb.save(file_path)

            # Success Message
            print(f"Excel file successfully saved as {file_path}")
            print('...')
        except Exception as e:
            print('!Error:', e, 345)
            print('...')
            return


from lib_sanitizer import format_month


def adjust_color(size):
    max_n = 16
    factor = 1 - (size / max_n)  # n이 증가할수록 색상이 더 진해짐
    """주어진 색상과 진하기 정도를 받아 새로운 색상을 반환"""
    base_color = (255, 255, 255)  # 흰색
    return tuple(int(c * factor) for c in base_color)


class ExpandTreeGenerator:
    def __init__(self, branch: Branch, period: list):
        self.period = []
        self.make_period(period)

        self.root = branch
        self.account_book = {}
        self.dfs_make_list()

    def make_period(self, period):
        begin, end = format_month(period[0]), format_month(period[1])
        cur_time = begin
        while cur_time <= end:
            self.period.append(cur_time)
            yy, mm = map(int, cur_time.split('-'))
            mm += 1
            if mm > 12:
                yy, mm = yy + 1, 1
            cur_time = format_month(f'{yy}-{mm}')

    def dfs_make_list(self):
        queue = deque([self.root])
        while queue:
            node = queue.pop()
            self.account_book[node.path] = {t: [0, 0] for t in self.period}

            for child_node in list(node.children.values())[::-1]:
                queue.append(child_node)

    def make_account_book(self, transactions):
        for _date, _branch, _cashflow, _description in transactions:
            _date = _date[:7:]

            while len(self.root.path) <= len(_branch):
                if _branch in self.account_book:
                    if _cashflow > 0:
                        self.account_book[_branch][_date][0] += _cashflow
                    elif _cashflow < 0:
                        self.account_book[_branch][_date][1] += _cashflow

                # Parent Node
                u = _branch.split('/')
                _branch = '/'.join(u[:len(u) - 1:])

        key_box = list(self.account_book.keys())
        for _branch in key_box:
            in_sum = sum([a[0] for a in self.account_book[_branch].values()])
            out_sum = sum([a[1] for a in self.account_book[_branch].values()])
            self.account_book[_branch]['Sum'] = [in_sum, out_sum]
        self.period.append('Sum')

    def make_excel(self, file_path_in, file_path_out):
        try:
            data_in = {branch: [self.account_book[branch][date][0] for date in self.period] for branch in self.account_book}
            data_out = {branch: [self.account_book[branch][date][1] for date in self.period] for branch in self.account_book}

            df_in = pd.DataFrame(data_in, index=self.period).transpose()
            df_out = pd.DataFrame(data_out, index=self.period).transpose()

            self._save_to_excel(df_in, file_path_in, "In")
            self._save_to_excel(df_out, file_path_out, "Out")
        except Exception as e:
            print('Error:', e)

    def _save_to_excel(self, df: pd.DataFrame, file_path, sheet_name):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Adding DataFrame to worksheet
        header_fill = PatternFill(start_color='012345', end_color="012345", fill_type="solid")
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True)):
            if r_idx == 1:
                continue
            r_idx = r_idx - 1 if r_idx > 1 else 0

            box_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            if row[0]:
                depth = len(row[0].split('/'))
                color = adjust_color(depth)
                hex_color = '{:02X}{:02X}{:02X}'.format(*color)
                box_fill = PatternFill(start_color=hex_color, end_color="ABCDEF", fill_type="solid")

            for c_idx, value in enumerate(row):
                if type(value) == int:
                    value = format_cost(value)

                cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                if c_idx == 0:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Styling headers
                cell.fill = box_fill
                if r_idx == 0:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color='FEFEFE')
                elif c_idx == 0:
                    cell.font = Font(bold=True, size=7.5)
                else:
                    cell.font = Font(size=8)

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
